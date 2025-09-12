# Automatically append new bets to Bet Log
# Fill in Bet Type, Selection, Stake, Odds, Decimal Odds, Result
# Set Stake = 0 for bonus bets
# Keep formulas for Payout ($), Net PnL ($), and Cumulative PnL ($) intact
# Avoid overwriting existing data

import pandas as pd
import openpyxl

FILE_PATH = "Bet_Tracker.xlsx"

def american_to_decimal(odds):
    """Convert American odds to Decimal odds."""
    if pd.isna(odds):
        return 0
    return 1 + (odds / 100 if odds > 0 else 100 / abs(odds))

def add_bet_to_excel(date, sportsbook, bet_type, selection, odds, stake=0, result="", bonus=False):
    """
    Append a new bet to the Excel Bet Log.

    Parameters:
    - date: str, e.g., "9/12/25"
    - sportsbook: str
    - bet_type: str
    - selection: str
    - odds: int (American odds, e.g., -125, +150)
    - stake: float (cash bet; ignored if bonus=True)
    - result: str ("Win", "Loss", or leave blank if pending)
    - bonus: bool (True if this is a bonus bet, stake treated as $0)
    """
    # Load or create Bet Log
    try:
        wb = openpyxl.load_workbook(FILE_PATH)
        ws = wb["Bet Log"]
    except FileNotFoundError:
        # Create new workbook with headers
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Bet Log"
        headers = ["Date", "Sportsbook", "Bet Type", "Selection", "Stake ($)", "Odds", "Result", "Decimal Odds", "Payout ($)", "Net PnL ($)", "Cumulative PnL ($)"]
        ws.append(headers)

    # Determine row to append
    next_row = ws.max_row + 1

    # Handle bonus stake
    actual_stake = 0 if bonus else stake

    # Decimal odds
    dec_odds = american_to_decimal(odds)

    # Append row (formulas for Payout, Net PnL, Cumulative PnL)
    ws.cell(row=next_row, column=1, value=date)
    ws.cell(row=next_row, column=2, value=sportsbook)
    ws.cell(row=next_row, column=3, value=bet_type)
    ws.cell(row=next_row, column=4, value=selection)
    ws.cell(row=next_row, column=5, value=actual_stake)
    ws.cell(row=next_row, column=6, value=odds)
    ws.cell(row=next_row, column=7, value=result)
    ws.cell(row=next_row, column=8, value=dec_odds)
    
    # Formulas
    ws.cell(row=next_row, column=9, value=f'=IF(G{next_row}="Win", E{next_row}*H{next_row}, 0)')
    ws.cell(row=next_row, column=10, value=f'=I{next_row}-E{next_row}')
    ws.cell(row=next_row, column=11, value=f'=SUM($J$2:J{next_row})')

    # Save workbook
    wb.save(FILE_PATH)
    print(f"Added bet to row {next_row}: {bet_type} - {selection} ({sportsbook})")