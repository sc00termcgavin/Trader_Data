import openpyxl
from openpyxl.chart import LineChart, BarChart, Reference
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import CellIsRule
from datetime import datetime
from openpyxl.chart.marker import DataPoint
from openpyxl.utils import get_column_letter

FILE_PATH = "Bet_Tracker.xlsx"

HEADERS = [
    "Date",
    "Sportsbook",
    "League",
    "Market",
    "Pick",
    "Stake ($)",
    "Odds",
    "Result",
    "Bonus",
    "Decimal Odds",
    "Payout ($)",
    "Net PnL ($)",
    "Cumulative PnL ($)"
]


def ensure_bet_log_headers(ws):
    header_values = [cell.value for cell in ws[1]] if ws.max_row >= 1 else []

    if not any(header_values):
        ws.append(HEADERS)
        return

    rename_map = {"Bet Type": "Market", "Selection": "Pick"}
    for idx, value in enumerate(header_values, start=1):
        if value in rename_map:
            ws.cell(row=1, column=idx, value=rename_map[value])

    header_values = [cell.value for cell in ws[1]]

    if "League" not in header_values:
        try:
            sportsbook_idx = header_values.index("Sportsbook") + 1
        except ValueError:
            sportsbook_idx = 2
        ws.insert_cols(sportsbook_idx + 1)
        ws.cell(row=1, column=sportsbook_idx + 1, value="League")

    for idx, header in enumerate(HEADERS, start=1):
        ws.cell(row=1, column=idx, value=header)


STAKE_COL = HEADERS.index("Stake ($)") + 1
RESULT_COL = HEADERS.index("Result") + 1
NET_PNL_COL = HEADERS.index("Net PnL ($)") + 1
CUM_PNL_COL = HEADERS.index("Cumulative PnL ($)") + 1

STAKE_COL_LETTER = get_column_letter(STAKE_COL)
RESULT_COL_LETTER = get_column_letter(RESULT_COL)
NET_PNL_COL_LETTER = get_column_letter(NET_PNL_COL)
CUM_PNL_COL_LETTER = get_column_letter(CUM_PNL_COL)

# -------------------------------
# 1. Load or create Bet Log
# -------------------------------
try:
    wb = openpyxl.load_workbook(FILE_PATH)
    ws_log = wb["Bet Log"]
except FileNotFoundError:
    wb = openpyxl.Workbook()
    ws_log = wb.active
    ws_log.title = "Bet Log"
    ws_log.append(HEADERS)
    wb.save(FILE_PATH)
else:
    ensure_bet_log_headers(ws_log)

# -------------------------------
# 1a. Convert Date column to Excel datetime format
# -------------------------------
for row in range(2, ws_log.max_row + 1):
    cell = ws_log[f"A{row}"]
    try:
        if isinstance(cell.value, str):
            cell.value = datetime.strptime(cell.value, "%m/%d/%y")
            cell.number_format = "mm/dd/yy"
    except Exception:
        pass

# -------------------------------
# 2. Conditional formatting for Net PnL
# -------------------------------
if ws_log.max_row > 1:
    net_pnl_range = f"{NET_PNL_COL_LETTER}2:{NET_PNL_COL_LETTER}{ws_log.max_row}"
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    ws_log.conditional_formatting.add(
        net_pnl_range,
        CellIsRule(operator='greaterThan', formula=['0'], fill=green_fill)
    )
    ws_log.conditional_formatting.add(
        net_pnl_range,
        CellIsRule(operator='lessThan', formula=['0'], fill=red_fill)
    )

# -------------------------------
# 3. Create or refresh Dashboard sheet
# -------------------------------
if "Dashboard" in wb.sheetnames:
    del wb["Dashboard"]
ws_dash = wb.create_sheet("Dashboard")

# -------------------------------
# 4. Dynamic KPI Section
# -------------------------------
ws_dash["A1"], ws_dash["B1"] = "KPI", "Value"
ws_dash["A2"], ws_dash["B2"] = "Total PnL ($)", (
    f"=SUM('Bet Log'!{NET_PNL_COL_LETTER}2:{NET_PNL_COL_LETTER}{ws_log.max_row})"
)
ws_dash["A3"], ws_dash["B3"] = "Total Stake ($)", (
    f"=SUM('Bet Log'!{STAKE_COL_LETTER}2:{STAKE_COL_LETTER}{ws_log.max_row})"
)
ws_dash["A4"], ws_dash["B4"] = "Wins", (
    f"=COUNTIF('Bet Log'!{RESULT_COL_LETTER}2:{RESULT_COL_LETTER}{ws_log.max_row},\"Win\")"
)
ws_dash["A5"], ws_dash["B5"] = "Total Bets", (
    f"=COUNTA('Bet Log'!{RESULT_COL_LETTER}2:{RESULT_COL_LETTER}{ws_log.max_row})"
)
ws_dash["A6"], ws_dash["B6"] = "Pending Bets", (
    f"=COUNTIF('Bet Log'!{RESULT_COL_LETTER}2:{RESULT_COL_LETTER}{ws_log.max_row},\"\")"
)
ws_dash["A7"], ws_dash["B7"] = "Win %", f"=IF(B5=0,0,B4/B5)"
ws_dash["A8"], ws_dash["B8"] = "ROI (%)", f"=IF(B3=0,0,B2/B3)"

# -------------------------------
# 5. Charts
# -------------------------------
# Line Chart: Cumulative Net PnL
line = LineChart()
line.title = "Cumulative Net PnL Over Time"
line.x_axis.title = "Date"
line.y_axis.title = "Cumulative Net PnL ($)"

dates = Reference(ws_log, min_col=1, min_row=2, max_row=ws_log.max_row)
cumulative = Reference(ws_log, min_col=CUM_PNL_COL, min_row=2, max_row=ws_log.max_row)
line.add_data(cumulative, titles_from_data=False)
line.set_categories(dates)

line.series[0].graphicalProperties.line.width = 20000  # thicker line
line.series[0].graphicalProperties.line.solidFill = "00B050"  # green
line.height = 10
line.width = 20

line.legend = None
ws_dash.add_chart(line, "A12")

# Bar Chart: Net PnL by Sportsbook
bar = BarChart()
bar.title = "Net PnL by Sportsbook"
bar.x_axis.title = "Sportsbook"
bar.y_axis.title = "Net PnL ($)"
sportsbooks = Reference(ws_log, min_col=2, min_row=2, max_row=ws_log.max_row)
net_pnl = Reference(ws_log, min_col=NET_PNL_COL, min_row=2, max_row=ws_log.max_row)
bar.add_data(net_pnl, titles_from_data=False)
bar.set_categories(sportsbooks)
bar.height, bar.width = 10, 20
ws_dash.add_chart(bar, "L12")

# -------------------------------
# 6. Save workbook
# -------------------------------
wb.save(FILE_PATH)
print(f"✅ Bet Tracker Dashboard ready: {FILE_PATH}")
