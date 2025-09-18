import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import CellIsRule
from datetime import datetime
import os

FILE_PATH = "Bet_Tracker.xlsx"

HEADERS = [
    "Date",
    "Sportsbook",
    "League",
    "Market",
    "Pick",
    "Stake ($)",
    "Odds",
    "Result",
    "Bonus",
    "Decimal Odds",
    "Payout ($)",
    "Net PnL ($)",
    "Cumulative PnL ($)"
]

# -------------------------------
# Utility: Convert American odds to Decimal
# -------------------------------
def american_to_decimal(odds):
    if odds == 0 or odds is None:
        return 0
    return 1 + (odds / 100 if odds > 0 else 100 / abs(odds))

# -------------------------------
# Header maintenance
# -------------------------------
def ensure_bet_log_headers(ws):
    header_values = [cell.value for cell in ws[1]] if ws.max_row >= 1 else []

    if not any(header_values):
        ws.append(HEADERS)
        return

    rename_map = {"Bet Type": "Market", "Selection": "Pick"}
    for idx, value in enumerate(header_values, start=1):
        if value in rename_map:
            ws.cell(row=1, column=idx, value=rename_map[value])

    header_values = [cell.value for cell in ws[1]]

    if "League" not in header_values:
        try:
            sportsbook_idx = header_values.index("Sportsbook") + 1
        except ValueError:
            sportsbook_idx = 2
        ws.insert_cols(sportsbook_idx + 1)
        ws.cell(row=1, column=sportsbook_idx + 1, value="League")

    for idx, header in enumerate(HEADERS, start=1):
        ws.cell(row=1, column=idx, value=header)


# -------------------------------
# Log a single bet
# -------------------------------
def log_bet(date, sportsbook, league, market, pick, odds, stake=0, result="", bonus=False):
    """
    Append a single bet to Bet Tracker.
    Bonus bets: stake is 0, payout = real money won.
    """
    # Load workbook or create if not exist
    try:
        wb = openpyxl.load_workbook(FILE_PATH)
        ws = wb["Bet Log"]
    except FileNotFoundError:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Bet Log"
        ws.append(HEADERS)
    else:
        ensure_bet_log_headers(ws)

    next_row = ws.max_row + 1

    # Actual stake for bonus bets
    actual_stake = 0 if bonus else stake
    dec_odds = american_to_decimal(odds)

    # Append data
    ws[f"A{next_row}"] = date
    ws[f"B{next_row}"] = sportsbook
    ws[f"C{next_row}"] = league
    ws[f"D{next_row}"] = market
    ws[f"E{next_row}"] = pick
    ws[f"F{next_row}"] = actual_stake
    ws[f"G{next_row}"] = odds
    ws[f"H{next_row}"] = result
    ws[f"I{next_row}"] = bonus
    ws[f"J{next_row}"] = dec_odds

    # -------------------------------
    # Payout Formula (handles Win, Push, Loss, Bonus)
    # -------------------------------
    if bonus:
        ws[f"K{next_row}"] = f'=IF(H{next_row}="Win",{stake}*(J{next_row}-1),IF(H{next_row}="Push",F{next_row},0))'
    else:
        ws[f"K{next_row}"] = f'=IF(H{next_row}="Win",F{next_row}*J{next_row},IF(H{next_row}="Push",F{next_row},0))'

    # Net PnL
    ws[f"L{next_row}"] = f'=IF(H{next_row}="", "", K{next_row}-F{next_row})'

    # Cumulative PnL
    ws[f"M{next_row}"] = f'=IF(L{next_row}="", "", SUM(L$2:L{next_row}))'

    # Conditional formatting for Net PnL
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    ws.conditional_formatting.add(f"L2:L{ws.max_row}",
                                  CellIsRule(operator='greaterThan', formula=['0'], fill=green_fill))
    ws.conditional_formatting.add(f"L2:L{ws.max_row}",
                                  CellIsRule(operator='lessThan', formula=['0'], fill=red_fill))

    # -------------------------------
    # Dashboard - update KPIs
    # -------------------------------
    if "Dashboard" in wb.sheetnames:
        ws_dash = wb["Dashboard"]
        # Clear old values (optional)
        for row in ws_dash.iter_rows(min_row=2, max_col=2, max_row=20):
            for cell in row:
                cell.value = None
    else:
        ws_dash = wb.create_sheet("Dashboard")
        ws_dash["A1"], ws_dash["B1"] = "Metric", "Value"

    # Metrics
    ws_dash["A2"], ws_dash["B2"] = "Total PnL ($)", f"=SUM('Bet Log'!L2:L{ws.max_row})"
    ws_dash["A3"], ws_dash["B3"] = "Total Stake ($)", f"=SUM('Bet Log'!F2:F{ws.max_row})"
    ws_dash["A4"], ws_dash["B4"] = "Wins", f'=COUNTIF(\'Bet Log\'!H2:H{ws.max_row},"Win")'
    ws_dash["A5"], ws_dash["B5"] = "Total Bets", f'=COUNTA(\'Bet Log\'!H2:H{ws.max_row})'
    ws_dash["A6"], ws_dash["B6"] = "Pending Bets", f'=COUNTIF(\'Bet Log\'!H2:H{ws.max_row},"")'
    ws_dash["A7"], ws_dash["B7"] = "Win %", f"=IF(B5=0,0,B4/B5)"
    ws_dash["A8"], ws_dash["B8"] = "ROI (%)", f"=IF(B3=0,0,B2/B3)"

    wb.save(FILE_PATH)
    print(f"✅ Logged bet in row {next_row}: {league} {market} - {pick} ({sportsbook})")


# -------------------------------
# Interactive Mode
# -------------------------------
if __name__ == "__main__":
    print("📊 Bet Logger - Enter your bets (type 'done' at sportsbook to stop)\n")

    while True:
        sportsbook = input("Sportsbook (or 'done' to quit): ").strip()
        if sportsbook.lower() == "done":
            break

        # Default date = today
        date = input(f"Date (MM/DD/YY, default today {datetime.today().strftime('%m/%d/%y')}): ").strip()
        if date == "":
            date = datetime.today().strftime("%m/%d/%y")

        league = input("League (e.g., NFL, NBA): ").strip()
        market = input("Market: ").strip()
        pick = input("Pick: ").strip()
        odds = int(input("Odds (e.g., -110, +125): ").strip())
        stake = float(input("Stake ($): ").strip())
        result = input("Result (Win/Loss/Push/blank): ").strip()
        bonus = input("Bonus bet? (y/n): ").strip().lower() == "y"

        log_bet(date, sportsbook, league, market, pick, odds, stake, result, bonus)

    print("✅ All bets logged successfully!")


# # -------------------------------
# # Bets
# # -------------------------------

# # 1. Fanatics $10 moneyline loss gave $100 bonus bets
# log_bet(
#     date="9/5/25",
#     sportsbook="Fanatics",
#     league="NFL",
#     market="Moneyline",
#     pick="Kansas City Chiefs ML vs LA Chargers",
#     odds=-170,
#     stake=10,
#     result="Loss",
#     bonus=True
# )


# # 3. Hard Rock $80 EVEN loss (+125)
# log_bet(
#     date="9/11/25",
#     sportsbook="Hard Rock",
#     league="NFL",
#     market="Total Points Odd/Even",
#     pick="Washington Commanders vs Green Bay Packers - Even",
#     odds=125,
#     stake=80,
#     result="Loss",
#     bonus=False
# )

# # 4. Fanatics $100 bonus bet on Total Points ODD (-125) -> won $80
# log_bet(
#     date="9/11/25",
#     sportsbook="Fanatics",
#     league="NFL",
#     market="Total Points Odd/Even",
#     pick="Washington Commanders vs Green Bay Packers - Odd",
#     odds=-125,
#     stake=100,  # real money won from bonus
#     result="Win",
#     bonus=True
# )

# # 5. Hard Rock Bets $25 bonus bet on 4-leg parlay (+364) -> potential win of 91
# log_bet(
#     date="9/12/25",
#     sportsbook="Hard Rock",
#     league="NFL",
#     market="4-leg parlay, Bills, Ravens, Bangels, Chiefs = win",
#     pick="Bills, Ravens, Bangels, Chiefs = win",
#     odds=+364,
#     stake=25,  # real money won from bonus
#     result="",
#     bonus=True
# )

# # 6. Hard Rock Bets $25 bonus bet on Atlanta spread of -5.54-leg parlay (+364) -> potential win of 93.7
# log_bet(
#     date="9/12/25",
#     sportsbook="Hard Rock",
#     league="NFL",
#     market="Point-Spread",
#     pick="Falcons -5.5",
#     odds=+375,
#     stake=25,  # real money won from bonus
#     result="",
#     bonus=True
# )

# #7 Hard Rock Bets $25 bonus bet on Jahmyr Gibbs to have 5+ Q1 Recieving Yards (+125) -> win of 31.25
# log_bet(
#     date="9/13/25",
#     sportsbook="Hard Rock",
#     league="NFL",
#     market="To have 5+ Q1 Recieving Yards",
#     pick="Jahmyr Gibbs",
#     odds=+125,
#     stake=25,  # real money won from bonus
#     result="",
#     bonus=True
# )

