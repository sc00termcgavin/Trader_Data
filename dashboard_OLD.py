import openpyxl
from openpyxl.chart import LineChart, BarChart, Reference

FILE_PATH = "Bet_Tracker.xlsx"

# -------------------------------
# 1. Create workbook if not exists
# -------------------------------
try:
    wb = openpyxl.load_workbook(FILE_PATH)
    ws_log = wb["Bet Log"]
except FileNotFoundError:
    wb = openpyxl.Workbook()
    ws_log = wb.active
    ws_log.title = "Bet Log"
    headers = [
        "Date", "Sportsbook", "Bet Type", "Selection",
        "Stake ($)", "Odds", "Result", "Bonus",
        "Decimal Odds", "Payout ($)", "Net PnL ($)", "Cumulative PnL ($)"
    ]
    ws_log.append(headers)
    wb.save(FILE_PATH)

# -------------------------------
# 2. Write formulas for existing rows
# -------------------------------
for row in range(2, ws_log.max_row + 1):
    stake = f"E{row}"
    odds = f"F{row}"
    result = f"G{row}"
    bonus = f"H{row}"
    dec_odds = f"I{row}"

    # Decimal Odds
    ws_log[dec_odds] = f'=IF(ISNUMBER({odds}),IF({odds}>0,1+{odds}/100,1+100/ABS({odds})),"")'
    # Payout
    ws_log[f"J{row}"] = f'=IF({result}="Win",IF({bonus}=TRUE,{stake}*({dec_odds}-1),{stake}*{dec_odds}),0)'
    # Net PnL
    ws_log[f"K{row}"] = f'=J{row}-{stake}'
    # Cumulative PnL
    ws_log[f"L{row}"] = f'=SUM(K$2:K{row})'

# -------------------------------
# 3. Create Dashboard Sheet
# -------------------------------
if "Dashboard" in wb.sheetnames:
    del wb["Dashboard"]
ws_dash = wb.create_sheet("Dashboard")

# KPI Section
ws_dash["A1"], ws_dash["B1"] = "KPI", "Value"
ws_dash["A2"], ws_dash["B2"] = "Total PnL ($)", "=SUM('Bet Log'!K2:K{})".format(ws_log.max_row)
ws_dash["A3"], ws_dash["B3"] = "Total Stake ($)", "=SUM('Bet Log'!E2:E{})".format(ws_log.max_row)
ws_dash["A4"], ws_dash["B4"] = "Win %", '=COUNTIF(\'Bet Log\'!G2:G{},"Win")/COUNTA(\'Bet Log\'!G2:G{})'.format(ws_log.max_row, ws_log.max_row)
ws_dash["A5"], ws_dash["B5"] = "ROI (%)", "=B2/B3"

# -------------------------------
# 4. Charts
# -------------------------------
# Line Chart: Cumulative Net PnL
line = LineChart()
line.title = "Cumulative Net PnL Over Time"
line.x_axis.title, line.y_axis.title = "Date", "Cumulative Net PnL ($)"
dates = Reference(ws_log, min_col=1, min_row=2, max_row=ws_log.max_row)
cumulative = Reference(ws_log, min_col=12, min_row=2, max_row=ws_log.max_row)
line.add_data(cumulative, titles_from_data=False)
line.set_categories(dates)
line.height, line.width = 10, 20
ws_dash.add_chart(line, "A8")

# Bar Chart: Net PnL by Sportsbook
bar = BarChart()
bar.title = "Net PnL by Sportsbook"
bar.x_axis.title, bar.y_axis.title = "Sportsbook", "Net PnL ($)"
sportsbooks = Reference(ws_log, min_col=2, min_row=2, max_row=ws_log.max_row)
net_pnl = Reference(ws_log, min_col=11, min_row=2, max_row=ws_log.max_row)
bar.add_data(net_pnl, titles_from_data=False)
bar.set_categories(sportsbooks)
bar.height, bar.width = 10, 20
ws_dash.add_chart(bar, "L8")

# -------------------------------
# 5. Save workbook
# -------------------------------
wb.save(FILE_PATH)
print(f"✅ Bet Tracker Dashboard ready: {FILE_PATH}")