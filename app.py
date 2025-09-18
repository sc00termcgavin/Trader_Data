import streamlit as st
from datetime import datetime
import openpyxl
from log_new_bets import log_bet  # reuse your existing function

FILE_PATH = "Bet_Tracker.xlsx"

st.set_page_config(page_title="📊 Bet Tracker", layout="wide")

st.title("📊 Bet Tracker Dashboard")

# -------------------------------
# Sidebar - Log New Bet
# -------------------------------
st.sidebar.header("➕ Log a New Bet")

with st.sidebar.form("log_bet_form"):
    sportsbook = st.text_input("Sportsbook")

    # League - required dropdown
    league = st.selectbox(
        "League",
        ["Select...", "NFL", "NBA", "MLB", "NHL", "EPL", "UFC", "Other"],
        index=0
    )

    # Market - required dropdown
    market = st.selectbox(
        "Market",
        ["Select...", "Moneyline", "Spread", "Total", "Prop", "Parlay"],
        index=0
    )

    pick = st.text_input("Pick / Wager")
    odds = st.number_input("Odds (American)", step=1, value=-110)
    stake = st.number_input("Stake ($)", step=1.0, value=10.0)
    result = st.selectbox("Result", ["Open", "Win", "Loss", "Push"])
    bonus = st.checkbox("Bonus Bet?")
    date = st.date_input("Date", datetime.today())

    submitted = st.form_submit_button("Log Bet")

    if submitted:
        # Validation
        if not sportsbook:
            st.error("❌ Sportsbook is required.")
        elif league == "Select...":
            st.error("❌ Please select a League.")
        elif market == "Select...":
            st.error("❌ Please select a Market.")
        elif not pick:
            st.error("❌ Pick / Wager is required.")
        else:
            log_bet(
                date.strftime("%m/%d/%y"),
                sportsbook,
                league,
                market,
                pick,
                odds,
                stake,
                result,
                bonus
            )
            st.success(f"✅ Logged bet: {league} - {market} - {pick} ({sportsbook})")

# -------------------------------
# Main Dashboard
# -------------------------------
try:
    wb = openpyxl.load_workbook(FILE_PATH, data_only=True)
    ws_log = wb["Bet Log"]

    st.subheader("📑 Bet Log")

    header_row = next(ws_log.iter_rows(min_row=1, max_row=1, values_only=True), None)
    headers = []
    if header_row:
        for idx, value in enumerate(header_row):
            headers.append(value if value not in (None, "") else f"Column {idx + 1}")

    max_col = len(headers) if headers else ws_log.max_column
    data_rows = list(ws_log.iter_rows(min_row=2, max_col=max_col, values_only=True))
    table_data = []

    if headers:
        for row in data_rows:
            if not any(cell not in (None, "") for cell in row):
                continue
            row_dict = {
                headers[idx]: (row[idx] if idx < len(row) else None)
                for idx in range(len(headers))
            }
            table_data.append(row_dict)

    wb.close()

    # ----- attach RowID (actual Excel row number) so edits/deletes map to Excel -----
    # table_data currently contains rows as dicts keyed by headers from the sheet.
    # Excel row 2 is the first data row, so RowID = index + 2.
    for i, r in enumerate(table_data):
        r["RowID"] = i + 2  # keep as int

    st.write("")  # spacing
    st.subheader("🗂️ Bet Log (editable helpers)")
    st.dataframe(table_data, width="stretch")

    # -------------------------
    # Helpers to recalc values
    # -------------------------
    def _to_float(x, default=0.0):
        try:
            return float(x)
        except (TypeError, ValueError):
            return default

    def american_to_decimal_local(odds):
        if odds in ("", None):
            return 0.0
        odds = float(odds)
        return 1 + (odds / 100 if odds > 0 else 100 / abs(odds))

    def recompute_row_values(ws, row):
        """Recompute Payout (K), Net PnL (L), Cumulative (M) for a given row based on current fields."""
        stake = _to_float(ws[f"F{row}"].value)                 # Stake ($)
        odds  = _to_float(ws[f"G{row}"].value)
        result = (ws[f"H{row}"].value or "").strip()
        normalized_result = result.lower() if isinstance(result, str) else ""
        bonus = bool(ws[f"I{row}"].value)
        dec_odds = american_to_decimal_local(odds)
        ws[f"J{row}"] = dec_odds

        # payout / net
        if normalized_result == "open":
            payout = None
            net = None
        elif result:
            if bonus:
                if result == "Win":
                    payout = _to_float(ws[f"F{row}"].value) * (dec_odds - 1)
                elif result == "Push":
                    payout = 0.0  # if you prefer push on bonus returns stake, change to _to_float(ws[f"F{row}"].value)
                else:
                    payout = 0.0
            else:
                if result == "Win":
                    payout = stake * dec_odds
                elif result == "Push":
                    payout = stake
                else:
                    payout = 0.0
            net = payout - (0.0 if bonus else stake)
        else:
            payout = None
            net = None

        ws[f"K{row}"] = payout
        ws[f"L{row}"] = net
        # cumulative will be set after we recalc down the column

    def recompute_cumulative(ws):
        """Recompute cumulative PnL (M) from row 2..max."""
        last = 0.0
        for r in range(2, ws.max_row + 1):
            net = ws[f"L{r}"].value
            net = _to_float(net, default=0.0) if net not in ("", None) else None
            if net is None:
                ws[f"M{r}"] = None
            else:
                last = (last if isinstance(last, (int, float)) else 0.0) + net
                ws[f"M{r}"] = last

    # -------------------------
    # Edit a single row
    # -------------------------
    with st.expander("✏️ Edit a row"):
        if table_data:
            row_ids = [r["RowID"] for r in table_data]
            chosen = st.selectbox("Select RowID to edit", row_ids)
            current = next((r for r in table_data if r["RowID"] == chosen), None)

            if current:
                from datetime import datetime as _dt
                date_val = current.get("Date") or _dt.today().strftime("%m/%d/%y")
                sportsbook_val = current.get("Sportsbook") or ""
                league_val = current.get("League") or "NFL"
                market_val = current.get("Market") or "Moneyline"
                pick_val = current.get("Pick") or ""
                odds_val = _to_float(current.get("Odds"), -110)
                stake_val = _to_float(current.get("Stake ($)"), 10.0)
                result_val = (current.get("Result") or "")
                bonus_val = bool(current.get("Bonus"))

                leagues = ["NFL","NBA","MLB","NHL","EPL","UFC","Other"]
                markets = ["Moneyline","Spread","Total","Prop","Parlay"]

                c1, c2 = st.columns(2)
                with c1:
                    date_in = st.text_input("Date (MM/DD/YY)", value=str(date_val))
                    sportsbook_in = st.text_input("Sportsbook", value=sportsbook_val)
                    league_in = st.selectbox("League", leagues, index=leagues.index(league_val) if league_val in leagues else 0)
                    market_in = st.selectbox("Market", markets, index=markets.index(market_val) if market_val in markets else 0)
                with c2:
                    pick_in = st.text_input("Pick / Wager", value=pick_val)
                    odds_in = st.number_input("Odds (American)", step=1.0, value=float(odds_val))
                    stake_in = st.number_input("Stake ($)", step=1.0, value=float(stake_val))
                    result_in = st.selectbox("Result", ["Open", "Win", "Loss", "Push"],
                                             index=["Open","Win","Loss","Push"].index(result_val) if result_val in ["Open","Win","Loss","Push"] else 0)
                    bonus_in = st.checkbox("Bonus Bet?", value=bonus_val)

                if st.button("Save changes"):
                    import openpyxl
                    wb_edit = openpyxl.load_workbook(FILE_PATH)
                    ws = wb_edit["Bet Log"]

                    r = int(chosen)
                    ws[f"A{r}"] = date_in
                    ws[f"B{r}"] = sportsbook_in
                    ws[f"C{r}"] = league_in
                    ws[f"D{r}"] = market_in
                    ws[f"E{r}"] = pick_in
                    ws[f"F{r}"] = float(stake_in)
                    ws[f"G{r}"] = float(odds_in)
                    ws[f"H{r}"] = result_in
                    ws[f"I{r}"] = bool(bonus_in)

                    recompute_row_values(ws, r)
                    recompute_cumulative(ws)

                    wb_edit.save(FILE_PATH)
                    wb_edit.close()
                    st.success(f"Row {r} updated.")
                    st.rerun()
        else:
            st.info("No rows to edit yet.")

    # -------------------------
    # Delete rows
    # -------------------------
    with st.expander("🗑️ Delete rows"):
        if table_data:
            row_ids = [r["RowID"] for r in table_data]
            to_delete = st.multiselect("Select RowID(s) to delete", row_ids)
            if to_delete and st.button("Confirm delete"):
                import openpyxl
                wb_del = openpyxl.load_workbook(FILE_PATH)
                ws = wb_del["Bet Log"]

                # Delete from bottom up to preserve indices
                for rid in sorted(to_delete, reverse=True):
                    ws.delete_rows(int(rid), 1)

                # Recompute cumulative after deletions
                recompute_cumulative(ws)

                wb_del.save(FILE_PATH)
                wb_del.close()
                st.success(f"Deleted rows: {sorted(to_delete)}")
                st.rerun()
        else:
            st.info("No rows to delete yet.")

    # ---- KPIs computed directly from Bet Log (no Dashboard dependency) ----
    def _to_number(x):
        try:
            return float(x)
        except (TypeError, ValueError):
            return 0.0

    st.subheader("📈 KPIs")

    # Build safe list of dict rows keyed by header names (already done above as table_data)
    rows = table_data  # alias for readability

    total_pnl = sum(_to_number(r.get("Net PnL ($)")) for r in rows)
    total_stake = sum(_to_number(r.get("Stake ($)")) for r in rows)

    result_vals = [(r.get("Result") or "").strip() for r in rows]

    wins = sum(1 for v in result_vals if v == "Win")
    total_bets = sum(1 for v in result_vals if v not in ("", None, "Open"))
    pending_bets = sum(1 for v in result_vals if v in ("", "Open"))

    win_pct = (wins / total_bets) if total_bets > 0 else 0.0
    roi_pct = (total_pnl / total_stake) if total_stake > 0 else 0.0

    col1, col2, col3, col4 = st.columns(4)
    with col1: st.metric("Total PnL ($)", f"{total_pnl:,.2f}")
    with col2: st.metric("Total Stake ($)", f"{total_stake:,.2f}")
    with col3: st.metric("Win %", f"{win_pct*100:.2f}%")
    with col4: st.metric("ROI (%)", f"{roi_pct*100:.2f}%")

except FileNotFoundError:
    st.warning("⚠️ No Bet Tracker file found yet. Log your first bet to create one.")
